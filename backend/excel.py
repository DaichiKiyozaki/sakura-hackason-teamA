"""面接記録からExcelファイルを生成する。"""

from __future__ import annotations

from io import BytesIO
from typing import Any, Mapping

import pandas as pd
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side


# 見出しや罫線に共通して使うピンク系のスタイル。
_TITLE_FILL = PatternFill("solid", fgColor="F4B6C2")
_SECTION_FILL = PatternFill("solid", fgColor="F7CAD0")
_HEADER_FILL = PatternFill("solid", fgColor="FCE4EC")
_THIN_BORDER = Border(
    left=Side(style="thin", color="D89AAA"),
    right=Side(style="thin", color="D89AAA"),
    top=Side(style="thin", color="D89AAA"),
    bottom=Side(style="thin", color="D89AAA"),
)


def create_interview_excel(interview: Mapping[str, Any]) -> BytesIO:
    """APIへ入力された面接1件を、保存せず1シートのExcelに変換する。

    戻り値はメモリ上のExcelデータであり、ファイルやDBには書き込まない。
    """

    # 表形式の出力データはpandasのDataFrameとして組み立てる。
    basic_information = pd.DataFrame(
        (
            ("面接日", interview["interview_date"]),
            ("候補者名", interview["candidate_name"]),
            ("面接官名", interview["interviewer_name"]),
            ("全体の所感", interview.get("overall_comment", "")),
        ),
        columns=("項目", "内容"),
    )
    question_answers = pd.DataFrame.from_records(
        (
            {"質問": item["question"], "回答（要約）": item["answer_summary"]}
            for item in interview.get("question_answers", [])
        ),
        columns=("質問", "回答（要約）"),
    )

    # pandasからopenpyxlエンジンへ渡し、APIで返せるBytesIOへ直接書き込む。
    output = BytesIO()
    basic_start_row = 3
    question_section_row = basic_start_row + len(basic_information) + 1
    question_table_start_row = question_section_row + 1

    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        basic_information.to_excel(
            writer,
            sheet_name="面接記録",
            startrow=basic_start_row,
            header=False,
            index=False,
        )
        question_answers.to_excel(
            writer,
            sheet_name="面接記録",
            startrow=question_table_start_row,
            index=False,
        )

        # セル結合や色など、DataFrameで表現しない見た目はopenpyxlで整える。
        sheet = writer.sheets["面接記録"]
        sheet.sheet_view.showGridLines = False

        sheet.merge_cells("A1:B1")
        sheet["A1"] = "面接記録"
        sheet["A1"].font = Font(size=16, bold=True)
        sheet["A1"].fill = _TITLE_FILL
        sheet["A1"].alignment = Alignment(horizontal="center")

        _write_section_title(sheet, 3, "① 基本情報")
        _write_section_title(sheet, question_section_row + 1, "② 質問・回答")

        basic_end_row = basic_start_row + len(basic_information)
        for row in range(basic_start_row + 1, basic_end_row + 1):
            _style_label(sheet.cell(row=row, column=1))
            _style_value(sheet.cell(row=row, column=2))

        question_header_row = question_table_start_row + 1
        _style_table_header(sheet, question_header_row, column_count=2)
        _style_table_rows(
            sheet,
            start_row=question_header_row + 1,
            row_count=len(question_answers),
            column_count=2,
        )

        # 長い質問や回答が読みやすくなるように列幅を調整する。
        sheet.column_dimensions["A"].width = 42
        sheet.column_dimensions["B"].width = 70

    output.seek(0)
    return output


def create_excel_filename(candidate_name: str, interview_date: str) -> str:
    """候補者名と面接日からExcelファイル名を返す。"""

    date_text = interview_date.replace("-", "")
    return f"面接記録_{candidate_name}_{date_text}.xlsx"


def _write_section_title(sheet: Any, row: int, title: str) -> int:
    """2列を結合した「① 基本情報」などの区画見出しを書く。"""

    sheet.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
    cell = sheet.cell(row=row, column=1, value=title)
    cell.font = Font(bold=True)
    cell.fill = _SECTION_FILL
    cell.alignment = Alignment(vertical="center")
    return row + 1


def _style_table_header(sheet: Any, row: int, column_count: int) -> None:
    """質問・回答表の列見出しへスタイルを設定する。"""

    for column in range(1, column_count + 1):
        cell = sheet.cell(row=row, column=column)
        cell.font = Font(bold=True)
        cell.fill = _HEADER_FILL
        cell.border = _THIN_BORDER
        cell.alignment = Alignment(horizontal="center", vertical="center")


def _style_table_rows(
    sheet: Any,
    start_row: int,
    row_count: int,
    column_count: int,
) -> None:
    """pandasが書いた表のデータへ折り返しと罫線を設定する。"""

    for row in range(start_row, start_row + row_count):
        for column in range(1, column_count + 1):
            cell = sheet.cell(row=row, column=column)
            cell.border = _THIN_BORDER
            cell.alignment = Alignment(vertical="top", wrap_text=True)


def _style_label(cell: Any) -> None:
    """基本情報の左側にある項目名へスタイルを設定する。"""

    cell.font = Font(bold=True)
    cell.fill = _HEADER_FILL
    cell.border = _THIN_BORDER
    cell.alignment = Alignment(vertical="top")


def _style_value(cell: Any) -> None:
    """基本情報の右側にある値へスタイルを設定する。"""

    cell.border = _THIN_BORDER
    cell.alignment = Alignment(vertical="top", wrap_text=True)
